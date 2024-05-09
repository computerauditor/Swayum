import time
import pyautogui
import openpyxl 
import pyperclip
import tkinter as tk
from tkinter import messagebox

#-----------------WARING----------------------------
# Function to show user agreement dialog
def show_agreement():
    agreement_text = ("I agree to RUN this script on my own discretion, "
                      "the author of this script shall not be held responsible for:\n\n"
                      "1) Any damage this script causes.\n"
                      "2) Me getting fired because of this 🔥🔥.\n"
                      "3) ☢️ Thermonuclear War ☢️.\n"
                      "4) My Crush Rejecting Me 👀.\n"
                      "5) I won't touch this code until I know what I am doing.")
    if messagebox.askyesno("User Agreement", agreement_text):
        show_warning()
    else:
        root.destroy()

# Function to show warning message
def show_warning():
    warning_text = ("S.A.P.S.A.D.I ~ Made By Dev!! \n\n"
                    "1) I'AM AT THE SAP MENU.\n\n"
                    "2) I HAVE CLOSED ALL THE RUNNING EXCEL FILES.\n\n"
                    "3) I WON'T BE ABLE TO USE THIS PC 'UNLESS' THIS CODE FINISHES")
    messagebox.showwarning("Warning", warning_text)

# Create tkinter window
root = tk.Tk()
root.withdraw()

# Show user agreement dialog
show_agreement()
#-------------------------------------------------------------      ----
# Function to read doc numbers from a the .txt file
def read_doc_numbers(file_path):
    with open(file_path, 'r') as file:
        doc_numbers = file.readlines()
    # Remove whitespace and newline characters
    doc_numbers = [doc.strip() for doc in doc_numbers]
    return doc_numbers
#-----------------------------------------------------------------
sap_menu_icon_location = (146,77)
pyautogui.click(sap_menu_icon_location)
time.sleep(3) 
pyautogui.typewrite('/nMIR4')
pyautogui.press('ENTER')
#-----------------------------------------------------------------
excel_file = "sap_details.xlsx"
Invoice_doc_no = (381,240)
pyautogui.click(sap_menu_icon_location)
time.sleep(0.5) 
doc_numbers = read_doc_numbers('docnum.txt')

wb = openpyxl.load_workbook("sap_details.xlsx")
ws = wb["Invoice Tracker"]

try:
    for i, doc_no in enumerate(doc_numbers, start=2):
        pyautogui.typewrite(doc_no)  # Corrected variable name
        pyautogui.press('ENTER')

        # Locate the area on the screen where Invoice Date is displayed
        
        # INVOICE DATE
        pyautogui.moveTo(376, 369, 1)
        time.sleep(1)
        pyautogui.click(376, 369)
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(1)
        pyautogui.hotkey('ctrl', 'c')
        invoice_date = pyperclip.paste()
        time.sleep(0.2)
        
        # REFERENCE NUMBER
        pyautogui.click(688, 368)
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        invoice_number = pyperclip.paste()
        time.sleep(1)

        # INVOICE AMOUNT
        pyautogui.click(346, 446)
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        invoice_amount = pyperclip.paste()
        time.sleep(1)

        # # GO TO TDS SECTION
        # pyautogui.click(516, 303)
        # time.sleep(5)

        # VENDOR CODE
        pyautogui.click(316, 316) # GO TO DETAILS
        time.sleep(1)
        pyautogui.click(776, 440) # GO TO INV PARTY
        time.sleep(1)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        vendor_number = pyperclip.paste()
        time.sleep(1)

        # VENDOR CODE
        pyautogui.click(789, 807) # GO TO PO NUMBER
        time.sleep(2)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        PO_number = pyperclip.paste()
        time.sleep(1)
        
        # ITEM CODE
        pyautogui.doubleClick(789, 807)
        time.sleep(2)
        pyautogui.moveTo(380, 369)
        pyautogui.click(380, 369)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        item_Code = pyperclip.paste()

        # GRN QUANTITY

        pyautogui.moveTo(475,643)
        time.sleep(2)
        pyautogui.click(475,643)
        pyautogui.click(475,643)
        time.sleep(1)
        pyautogui.click(515,675)
        time.sleep(1)
        pyautogui.moveTo(229,379)
        pyautogui.click(229,379)
        time.sleep(1)
        pyautogui.click(384,639)
        time.sleep(1)
        pyautogui.click(406,741)
        time.sleep(8)
        pyautogui.click(714,657) # Excel Click in the middle to search previously(709,604)
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.typewrite('Tr./Ev. Goods receipt')
        pyautogui.press('ENTER')
        pyautogui.press('ENTER')
        for _ in range(10):
            pyautogui.press('TAB')
        pyautogui.press('ENTER')
        time.sleep(1)
        pyautogui.press('TAB')
        pyautogui.hotkey('ctrl', 'c')
        GRN_qty = pyperclip.paste()

        # INV_RECEIPT QUANTITY
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.typewrite('Tr./Ev. Invoice receipt')
        pyautogui.press('ENTER')
        pyautogui.press('ENTER')
        for _ in range(10):
            pyautogui.press('TAB')
        pyautogui.press('ENTER')
        time.sleep(1)
        pyautogui.press('TAB')
        pyautogui.hotkey('ctrl', 'c')
        InvoiceRec_qty = pyperclip.paste()    

        # PARK QUANTITY
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.typewrite('Tr./Ev. Parked invoice')
        pyautogui.press('ENTER')
        pyautogui.press('ENTER')
        for _ in range(10):
            pyautogui.press('TAB')
        pyautogui.press('ENTER')
        time.sleep(1)
        pyautogui.press('TAB')
        pyautogui.hotkey('ctrl', 'c')
        Park_qty = pyperclip.paste()   

        pyautogui.hotkey('alt', 'f4')
        pyautogui.press('TAB')
        time.sleep(1)
        pyautogui.press('ENTER')
        time.sleep(2)

        # Update Excel file
        ws[f'A{i}'] = doc_no
        ws[f'B{i}'] = invoice_date
        ws[f'C{i}'] = invoice_number
        ws[f'D{i}'] = vendor_number
        ws[f'E{i}'] = PO_number
        ws[f'F{i}'] = invoice_amount
        ws[f'G{i}'] = item_Code
        ws[f'H{i}'] = GRN_qty
        ws[f'I{i}'] = InvoiceRec_qty
        ws[f'J{i}'] = Park_qty
        
        # Save the Excel file
        wb.save(excel_file)

        pyautogui.moveTo(168, 79)
        time.sleep(0.5)
        pyautogui.click(168, 79)    
        time.sleep(1)
        pyautogui.click(168, 79)
        pyautogui.typewrite('/nMIR4')
        pyautogui.press('ENTER')
        time.sleep(8)

    messagebox.showinfo("Success", "Script completed successfully. Data saved to Excel file.")
except Exception as e:
    messagebox.showerror("Error", f"Script failed due to system error:\n{str(e)}")
